from __future__ import annotations

from dataclasses import dataclass
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from extract_eval import EvalExtractionResult
from extract_fs import FSExtractionResult
from parse_blank_vs_final import ManualInputCell
from utils import AuditRow, normalize_text


@dataclass
class MappingResult:
    writes: list[dict]
    missing: list[dict]
    audit_rows: list[AuditRow]


def _build_value_pool(fs: FSExtractionResult, ev: EvalExtractionResult) -> list[dict[str, Any]]:
    pool: list[dict[str, Any]] = []

    for _, row in fs.income_statement.iterrows():
        pool.append(
            {
                "key": f"fs:{row['canonical_line_item']}:y1",
                "label": str(row["line_item"]),
                "value": row["amount_year_1"],
                "source": "fs_pdf",
                "method": "line_item_parse",
                "confidence": 0.85,
            }
        )
        pool.append(
            {
                "key": f"fs:{row['canonical_line_item']}:y2",
                "label": str(row["line_item"]),
                "value": row["amount_year_2"],
                "source": "fs_pdf",
                "method": "line_item_parse",
                "confidence": 0.85,
            }
        )

    for _, row in ev.production.iterrows():
        pool.append(
            {
                "key": f"prod:month:{int(row['month'])}",
                "label": f"month {int(row['month'])} revenue",
                "value": row["gross_revenue"],
                "source": "eval_pdf",
                "method": "monthly_table_parse",
                "confidence": 0.8,
            }
        )
    if not ev.production.empty:
        pool.append(
            {
                "key": "prod:avg",
                "label": "average monthly revenue",
                "value": float(ev.production["gross_revenue"].mean()),
                "source": "eval_pdf",
                "method": "derived_average",
                "confidence": 0.7,
            }
        )

    for _, row in ev.provider_split.iterrows():
        pool.append(
            {
                "key": f"provider:{row['provider_type']}",
                "label": f"{row['provider_type']} percent",
                "value": row["percent"],
                "source": "eval_pdf",
                "method": "provider_split_parse",
                "confidence": 0.8,
            }
        )

    for _, row in ev.payroll.iterrows():
        if pd.notna(row.get("hourly_rate")):
            pool.append(
                {
                    "key": f"payroll:{row['role']}:hourly",
                    "label": f"{row['role']} hourly",
                    "value": row["hourly_rate"],
                    "source": "eval_pdf",
                    "method": "payroll_parse",
                    "confidence": 0.75,
                }
            )
        if pd.notna(row.get("inferred_hours_per_week")):
            pool.append(
                {
                    "key": f"payroll:{row['role']}:hours",
                    "label": f"{row['role']} hours",
                    "value": row["inferred_hours_per_week"],
                    "source": "eval_pdf",
                    "method": "payroll_parse_days_to_hours",
                    "confidence": 0.65,
                }
            )

    return pool


def _best_match(target_label: str, target_value: Any, pool: list[dict[str, Any]]) -> dict[str, Any] | None:
    # Priority 1: exact numeric match to final value.
    if isinstance(target_value, (int, float)):
        for item in pool:
            v = item.get("value")
            if isinstance(v, (int, float)) and abs(float(v) - float(target_value)) < 1e-6:
                return item

    # Priority 2: weak label overlap.
    norm_label = normalize_text(target_label)
    best = None
    best_score = 0
    for item in pool:
        label = normalize_text(str(item.get("label", "")))
        overlap = len(set(norm_label.split()) & set(label.split()))
        if overlap > best_score and overlap >= 1:
            best = item
            best_score = overlap
    return best


def _infer_label(ws, row: int, col: int) -> str:
    for offset in range(1, 5):
        c = ws.cell(row=row, column=max(1, col - offset)).value
        if isinstance(c, str) and c.strip():
            return c
    return ""


def map_and_fill(
    template_blank_path: str,
    output_path: str,
    manual_inputs: list[ManualInputCell],
    fs: FSExtractionResult,
    ev: EvalExtractionResult,
    deal_id: str,
    timestamp: str,
) -> MappingResult:
    wb = load_workbook(template_blank_path)
    pool = _build_value_pool(fs, ev)

    writes: list[dict] = []
    missing: list[dict] = []
    audit_rows: list[AuditRow] = []

    for target in manual_inputs:
        ws = wb[target.sheet]
        cell = ws[target.cell]
        label = _infer_label(ws, cell.row, cell.column)
        match = _best_match(label, target.final_value, pool)

        if match and match.get("value") is not None:
            cell.value = match["value"]
            writes.append({"sheet": target.sheet, "cell": target.cell, "value": match["value"]})
            audit_rows.append(
                AuditRow(
                    timestamp=timestamp,
                    deal_id=deal_id,
                    sheet=target.sheet,
                    cell=target.cell,
                    value_written=match["value"],
                    source_file=match["source"],
                    extraction_method=match["method"],
                    confidence=match["confidence"],
                    status="filled",
                    notes=f"matched key={match['key']} label={label}",
                )
            )
        else:
            missing.append(
                {
                    "sheet": target.sheet,
                    "cell": target.cell,
                    "expected": target.final_value,
                    "reason": "no_confident_match",
                }
            )
            audit_rows.append(
                AuditRow(
                    timestamp=timestamp,
                    deal_id=deal_id,
                    sheet=target.sheet,
                    cell=target.cell,
                    value_written="",
                    source_file="",
                    extraction_method="",
                    confidence=0.0,
                    status="missing",
                    notes="No confident extraction match",
                )
            )

    wb.save(output_path)
    return MappingResult(writes=writes, missing=missing, audit_rows=audit_rows)
