from __future__ import annotations

from dataclasses import dataclass
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from parse_blank_vs_final import ManualInputCell


@dataclass
class ValidationSummary:
    total: int
    matched: int
    mismatched: int
    matched_pct: float


def _delta(expected: Any, actual: Any) -> Any:
    if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
        return actual - expected
    return ""


def validate_output(
    output_path: str,
    final_path: str,
    manual_inputs: list[ManualInputCell],
) -> tuple[pd.DataFrame, ValidationSummary]:
    wb_out = load_workbook(output_path, data_only=True)
    wb_final = load_workbook(final_path, data_only=True)

    rows = []
    matched = 0

    for m in manual_inputs:
        if m.sheet not in wb_out.sheetnames or m.sheet not in wb_final.sheetnames:
            continue
        actual = wb_out[m.sheet][m.cell].value
        expected = wb_final[m.sheet][m.cell].value
        ok = actual == expected
        if ok:
            matched += 1
        rows.append(
            {
                "sheet": m.sheet,
                "cell": m.cell,
                "expected": expected,
                "actual": actual,
                "delta": _delta(expected, actual),
                "match": ok,
            }
        )

    df = pd.DataFrame(rows)
    total = len(df)
    mismatched = int((~df["match"]).sum()) if total else 0
    matched_pct = (matched / total * 100.0) if total else 0.0
    summary = ValidationSummary(total=total, matched=matched, mismatched=mismatched, matched_pct=matched_pct)
    return df, summary
