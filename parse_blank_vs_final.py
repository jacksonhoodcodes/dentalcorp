from __future__ import annotations

from dataclasses import dataclass
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


@dataclass
class ManualInputCell:
    sheet: str
    cell: str
    blank_value: object
    final_value: object


def _is_formula(cell: Cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def identify_manual_input_cells(blank_path: str, final_path: str) -> list[ManualInputCell]:
    """Diff blank vs final and return likely user-entered cells.

    Rules:
    - Ignore pure formula cells in either workbook.
    - Capture changed constants only.
    """
    wb_blank = load_workbook(blank_path, data_only=False)
    wb_final = load_workbook(final_path, data_only=False)

    manual_inputs: list[ManualInputCell] = []
    common_sheets = [s for s in wb_blank.sheetnames if s in wb_final.sheetnames]

    for sheet_name in common_sheets:
        ws_blank = wb_blank[sheet_name]
        ws_final = wb_final[sheet_name]

        max_row = max(ws_blank.max_row, ws_final.max_row)
        max_col = max(ws_blank.max_column, ws_final.max_column)

        for row in ws_blank.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for blank_cell in row:
                final_cell = ws_final.cell(blank_cell.row, blank_cell.column)

                if _is_formula(blank_cell) or _is_formula(final_cell):
                    continue

                blank_val = blank_cell.value
                final_val = final_cell.value

                if blank_val != final_val:
                    manual_inputs.append(
                        ManualInputCell(
                            sheet=sheet_name,
                            cell=blank_cell.coordinate,
                            blank_value=blank_val,
                            final_value=final_val,
                        )
                    )

    return manual_inputs


def manual_inputs_to_df(items: Iterable[ManualInputCell]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "sheet": i.sheet,
                "cell": i.cell,
                "blank_value": i.blank_value,
                "final_value": i.final_value,
            }
            for i in items
        ]
    )
